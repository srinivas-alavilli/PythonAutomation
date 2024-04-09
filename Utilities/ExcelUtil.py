import openpyxl

#To get the Row count in the Excel sheet

def getRowCount(file,sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    print("No of Rows: ", sheet.max_row)
    return (sheet.max_row)

#To get the Column count in the Excel sheet
def getColCount(file,sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    print("No of Columns: ", sheet.max_row)
    return (sheet.max_column)

#To Read the cell data for the given sheet, row and column
def readData(file,sheetName,rownum,colnum):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.cell(row=rownum, column=colnum).value

#To write the data in the given sheet, row and column
def writeData(file,sheetName,rownum, colnum, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    sheet.cell(row=rownum, column=colnum).value = data
    workbook.save(file)